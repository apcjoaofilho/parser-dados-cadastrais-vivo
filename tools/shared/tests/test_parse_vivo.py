import sys
from pathlib import Path

_parser_dir = Path(__file__).resolve().parent.parent.parent / 'vivo-parser'
sys.path.insert(0, str(_parser_dir))
_shared_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_shared_dir))

import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

from parse_vivo import (
    deduplicar_dataframe, estilizar_excel, limpar_valor,
    ordenar_dataframe, parse_arquivo, parse_registro,
    RE_CAMPO, RE_CPF_CONSULTA,
)
import geocoding


class TestRegexes:
    def test_cpf_consulta_match(self):
        assert RE_CPF_CONSULTA.match('* CPF: 069.552.203-57                                                          *')

    def test_cpf_consulta_no_match_registro(self):
        assert not RE_CPF_CONSULTA.match('* CPF:..........................................................084.221.183-78 *')

    def test_campo_situacao(self):
        m = RE_CAMPO.match('* SITUAÇÃO:............................................................INATIVO *')
        assert m
        assert m.group(1).strip() == 'SITUAÇÃO'
        assert limpar_valor(m.group(2)) == 'INATIVO'


class TestParseRegistro:
    def test_registro_completo(self):
        linhas = [
            '* NÚMERO DA LINHA:.............................................(11) 93257-9719 *',
            '* CLIENTE:................................FRANCISCO FABRICIO FERREIRA DA ROCHA *',
            '* CPF:..........................................................084.221.183-78 *',
            '* ENDEREÇO:...................................Rua R VILA PARANA 476 PX A PADAR *',
            '* BAIRRO:.............................................................BRASILIA *',
            '* CEP:..............................................................62.400-000 *',
            '* MUNICÍPIO:...........................................................CAMOCIM *',
            '* ESTADO:...................................................................CE *',
            '* MODALIDADE:..............................................................PRE *',
            '* SITUAÇÃO:............................................................INATIVO *',
            '* DATA HABILITAÇÃO:.................................................15/08/2024 *',
            '* DATA RESCISÃO:....................................................18/01/2025 *',
            '* ............................................................................ *',
        ]
        reg, nxt = parse_registro(linhas, 0, '084.221.183-78', 'teste.txt')
        assert reg is not None
        assert reg.numero_linha == '(11) 93257-9719'
        assert reg.cliente == 'FRANCISCO FABRICIO FERREIRA DA ROCHA'
        assert reg.situacao == 'INATIVO'
        assert reg.data_rescisao == '18/01/2025'

    def test_registro_sem_data_rescisao(self):
        linhas = [
            '* NÚMERO DA LINHA:.............................................(27) 99610-0091 *',
            '* CLIENTE:................................FRANCISCO FABRICIO FERREIRA DA ROCHA *',
            '* CPF:..........................................................084.221.183-78 *',
            '* ENDEREÇO:...................................Rua R VILA PARANA 476 PX A PADAR *',
            '* BAIRRO:.............................................................BRASILIA *',
            '* CEP:..............................................................62.400-000 *',
            '* MUNICÍPIO:...........................................................CAMOCIM *',
            '* ESTADO:...................................................................CE *',
            '* MODALIDADE:..............................................................PRE *',
            '* SITUAÇÃO:..............................................................ATIVO *',
            '* DATA HABILITAÇÃO:.................................................19/12/2025 *',
            '* ............................................................................ *',
        ]
        reg, nxt = parse_registro(linhas, 0, '084.221.183-78', 'teste.txt')
        assert reg is not None
        assert reg.situacao == 'ATIVO'
        assert reg.data_rescisao is None


class TestDeduplicacao:
    def test_remove_duplicados(self):
        df = pd.DataFrame({
            'cpf_consulta': ['123', '123'],
            'numero_linha': ['(11) 1111-1111', '(11) 1111-1111'],
            'cliente': ['JOAO', 'JOAO'],
            'cpf': ['123', '123'],
            'endereco': ['Rua A', 'Rua A'],
            'bairro': ['Centro', 'Centro'],
            'cep': ['01000-000', '01000-000'],
            'municipio': ['SP', 'SP'],
            'estado': ['SP', 'SP'],
            'modalidade': ['PRE', 'PRE'],
            'situacao': ['ATIVO', 'ATIVO'],
            'data_habilitacao': ['01/01/2024', '01/01/2024'],
            'data_rescisao': [None, None],
            'arquivo_origem': ['a.txt', 'b.txt'],
        })
        df_dedup = deduplicar_dataframe(df)
        assert len(df_dedup) == 1


class TestOrdenacao:
    def test_ativo_primeiro(self):
        df = pd.DataFrame({
            'cpf_consulta': ['1', '2'],
            'numero_linha': ['A', 'B'],
            'cliente': ['X', 'Y'],
            'cpf': ['1', '2'],
            'endereco': ['Rua', 'Rua'],
            'bairro': ['Centro', 'Centro'],
            'cep': ['00000-000', '00000-000'],
            'municipio': ['SP', 'SP'],
            'estado': ['SP', 'SP'],
            'modalidade': ['PRE', 'PRE'],
            'situacao': ['INATIVO', 'ATIVO'],
            'data_habilitacao': ['01/01/2024', '15/03/2025'],
            'data_rescisao': [None, None],
            'arquivo_origem': ['a.txt', 'b.txt'],
        })
        df_ord = ordenar_dataframe(df)
        assert df_ord.iloc[0]['situacao'] == 'ATIVO'


class TestEstilizacaoExcel:
    def test_estilo_ativo_verde(self):
        df = pd.DataFrame({
            'situacao': ['ATIVO', 'INATIVO'],
            'modalidade': ['PRE', 'POS'],
            'cliente': ['A', 'B'],
        })
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / 'test.xlsx'
            estilizar_excel(df, path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            ativo_fill = ws.cell(row=2, column=1).fill.start_color.rgb
            assert ativo_fill == '00C6EFCE' or ativo_fill == 'C6EFCE'

    def test_negrito_pos(self):
        df = pd.DataFrame({
            'situacao': ['INATIVO', 'INATIVO'],
            'modalidade': ['PRE', 'POS'],
            'cliente': ['A', 'B'],
        })
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / 'test.xlsx'
            estilizar_excel(df, path)
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            col_modalidade = df.columns.get_loc('modalidade') + 1
            assert ws.cell(row=3, column=col_modalidade).font.bold is True


class TestGeocoding:
    @patch('geocoding.requests.get')
    def test_geocodifica_e_cacheia(self, mock_get):
        mock_response = MagicMock()
        mock_response.json.return_value = [{'lat': '-23.5505', 'lon': '-46.6333'}]
        mock_response.raise_for_status = MagicMock()
        mock_get.return_value = mock_response

        df = pd.DataFrame({
            'endereco': ['Rua Augusta'],
            'bairro': ['Consolacao'],
            'municipio': ['Sao Paulo'],
            'estado': ['SP'],
            'cep': ['01305-000'],
        })
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / 'cache.db'
            geocoding.processar_geocoding(df, str(db_path))

        assert df.iloc[0]['latitude'] == -23.5505
        assert df.iloc[0]['longitude'] == -46.6333
        assert 'google.com/maps' in df.iloc[0]['google_maps_url']
