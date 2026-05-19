#!/usr/bin/env python3
"""
Testes unitários para parse_tim.py.
Cobertura: parsing de texto, filtros, snapshots, ZIP, geocodificação, Excel.
"""

import sys
from pathlib import Path

# Adiciona tools/tim-parser e tools/shared ao PYTHONPATH
_parser_dir = Path(__file__).resolve().parent.parent.parent / "tim-parser"
_shared_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_parser_dir))
sys.path.insert(0, str(_shared_dir))

import io
import tempfile
import zipfile
from unittest.mock import MagicMock, patch

import fitz
import pandas as pd
import pytest

from parse_tim import (
    CAMPOS_VALIDOS,
    LIXO_EXATO,
    MAX_COMPRESSION_RATIO,
    MAX_LINE_LENGTH,
    MAX_PDF_SIZE_BYTES,
    MAX_PDFS_PER_ZIP,
    MAX_ZIP_SIZE_BYTES,
    RE_CPF_ARQUIVO,
    RE_CPF_HEADER,
    RE_HEADER_FOOTER,
    RE_LABEL,
    RegistroTim,
    _parse_registro_tim,
    _sanitizar_nome_arquivo,
    _validar_caminho,
    _validar_tamanho_zip,
    extrair_cpf_do_nome_arquivo,
    extrair_cpf_do_texto,
    extrair_texto_pdf_de_bytes,
    filtrar_linhas,
    parse_texto_tim,
    processar_zip,
    selecionar_snapshots_recentes,
)
import geocoding
from pipeline import deduplicar_dataframe, estilizar_excel


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _criar_pdf_em_memoria(texto: str) -> bytes:
    """Cria um PDF simples em memória com o texto fornecido."""
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 50), texto)
    pdf_bytes = doc.tobytes()
    doc.close()
    return pdf_bytes


def _criar_zip_com_pdfs(pdfs: dict[str, bytes]) -> Path:
    """Cria um arquivo ZIP temporário com os PDFs fornecidos (nome -> bytes)."""
    tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
    with zipfile.ZipFile(tmp.name, "w") as zf:
        for nome, data in pdfs.items():
            zf.writestr(nome, data)
    return Path(tmp.name)


# ---------------------------------------------------------------------------
# Regexes
# ---------------------------------------------------------------------------

class TestRegexes:
    def test_cpf_header_match(self):
        texto = "Relatório de Cadastro por (CPF): 06319775386"
        m = RE_CPF_HEADER.search(texto)
        assert m
        assert m.group(1) == "06319775386"

    def test_cpf_header_no_match(self):
        assert not RE_CPF_HEADER.search("Relatório de Cadastro por CPF: 123")

    def test_cpf_arquivo_match(self):
        nome = "1_cadastro_84966_2026_7937535_06319775386_18052026012008953.pdf"
        m = RE_CPF_ARQUIVO.search(nome)
        assert m
        assert m.group(1) == "06319775386"

    def test_cpf_arquivo_no_match(self):
        assert not RE_CPF_ARQUIVO.search("relatorio.pdf")

    def test_label_match(self):
        assert RE_LABEL.match("NÚMERO DA LINHA:")
        assert RE_LABEL.match("CIDADE/UF - CEP RESIDENCIAL:")
        assert RE_LABEL.match("DATA INÍCIO VÍNCULO:")

    def test_label_no_match_valor(self):
        assert not RE_LABEL.match("5588997303482")
        assert not RE_LABEL.match("Ativo")

    def test_header_footer_match(self):
        assert RE_HEADER_FOOTER.match("1 /5")
        assert RE_HEADER_FOOTER.match(
            "Não foram encontrados registros referentes à informação solicitada."
        )
        assert RE_HEADER_FOOTER.match("Número Solicitação: 7937536")
        assert RE_HEADER_FOOTER.match(
            "Período de Pesquisa: 18/05/2026 00:00:00 até 18/05/2026 01:07:56"
        )

    def test_header_footer_no_match_campo(self):
        assert not RE_HEADER_FOOTER.match("NÚMERO DA LINHA:")
        assert not RE_HEADER_FOOTER.match("STATUS ATUAL:")
        assert not RE_HEADER_FOOTER.match("07/05/2026 17:46:04")


# ---------------------------------------------------------------------------
# Filtrar Linhas
# ---------------------------------------------------------------------------

class TestFiltrarLinhas:
    def test_remove_lixo(self):
        linhas = [
            "NÚMERO DA LINHA:",
            "5588",
            "aa",
            "a",
            "TIPO DA LINHA:",
            "PRÉ-PAGO",
            "a",
        ]
        filtradas = filtrar_linhas(linhas)
        assert filtradas == [
            "NÚMERO DA LINHA:",
            "5588",
            "TIPO DA LINHA:",
            "PRÉ-PAGO",
        ]

    def test_remove_headers(self):
        linhas = [
            "CONFIDENCIAL",
            "TIM S/A",
            "+55 11 4251-6633",
            "NÚMERO DA LINHA:",
            "5588",
        ]
        filtradas = filtrar_linhas(linhas)
        assert filtradas == ["NÚMERO DA LINHA:", "5588"]

    def test_mantem_valores_validos(self):
        linhas = ["STATUS ATUAL:", "Ativo", "DATA STATUS:", "07/05/2026 17:46:04"]
        filtradas = filtrar_linhas(linhas)
        assert filtradas == linhas

    def test_truncamento_reddos(self):
        linha_longa = "A" * (MAX_LINE_LENGTH + 100)
        filtradas = filtrar_linhas([linha_longa])
        assert len(filtradas[0]) == MAX_LINE_LENGTH


# ---------------------------------------------------------------------------
# Parse Texto TIM
# ---------------------------------------------------------------------------

class TestParseTextoTim:
    def test_registro_completo(self):
        texto = """NÚMERO DA LINHA:
5588997303482
aa
aa
TIPO DA LINHA:
PRÉ-PAGO
a
a
STATUS ATUAL:
Ativo
a
a
DATA STATUS:
08/05/2021 01:22:49
a
a
DATA INÍCIO VÍNCULO:
07/02/2020 10:45:37
a
a
DATA CADASTRO:
08/03/2020 10:12:37
a
a
DATA FIM VÍNCULO:
.
a
a
NOME:
MICKAEL JOSE MARQUES DE CARVALHO
a
a
TIPO DO CLIENTE:
CON
CPF/CNPJ:
063.197.753-86
SEXO:
MASCULINO
TIPO DOCUMENTO:
REGISTRO GERAL
DATA NASCIMENTO:
23/02/2000
NÚM. DOCUMENTO:
.
NACIONALIDADE:
.
DATA EMISSÃO:
.
TELEFONE CONTATO:
9936669873
PAÍS EMISSOR:
.
ENDEREÇO RESIDENCIAL:
PARA, SN, SAO PEDRO
CIDADE/UF - CEP RESIDENCIAL:
GRANJA/CE - 62430-000
ENDEREÇO FATURA:
.
CIDADE/UF - CEP FATURA:
.
"""
        regs = parse_texto_tim(texto, "06319775386", "teste.pdf")
        assert len(regs) == 1
        r = regs[0]
        assert r.cpf_consulta == "06319775386"
        assert r.numero_linha == "5588997303482"
        assert r.tipo_linha == "PRÉ-PAGO"
        assert r.status_atual == "Ativo"
        assert r.data_status == "08/05/2021 01:22:49"
        assert r.data_inicio_vinculo == "07/02/2020 10:45:37"
        assert r.data_cadastro == "08/03/2020 10:12:37"
        assert r.data_fim_vinculo is None
        assert r.nome == "MICKAEL JOSE MARQUES DE CARVALHO"
        assert r.tipo_cliente == "CON"
        assert r.cpf_cnpj == "063.197.753-86"
        assert r.sexo == "MASCULINO"
        assert r.tipo_documento == "REGISTRO GERAL"
        assert r.data_nascimento == "23/02/2000"
        assert r.num_documento is None
        assert r.nacionalidade == ""
        assert r.data_emissao is None
        assert r.telefone_contato == "9936669873"
        assert r.pais_emissor is None
        assert r.endereco_residencial == "PARA, SN, SAO PEDRO"
        assert r.cidade_uf_cep_residencial == "GRANJA/CE - 62430-000"
        assert r.endereco_fatura is None
        assert r.cidade_uf_cep_fatura is None
        assert r.arquivo_origem == "teste.pdf"

    def test_multiplos_registros(self):
        texto = """NÚMERO DA LINHA:
5588
TIPO DA LINHA:
PRÉ-PAGO
STATUS ATUAL:
Ativo
DATA CADASTRO:
01/01/2024 00:00:00
NOME:
JOAO
CPF/CNPJ:
123.456.789-00
SEXO:
MASCULINO
DATA NASCIMENTO:
01/01/1990
ENDEREÇO RESIDENCIAL:
RUA A
CIDADE/UF - CEP RESIDENCIAL:
SP/SP - 01000-000
NÚMERO DA LINHA:
5589
TIPO DA LINHA:
PÓS-PAGO
STATUS ATUAL:
Inativo
DATA CADASTRO:
02/02/2024 00:00:00
NOME:
MARIA
CPF/CNPJ:
987.654.321-00
SEXO:
FEMININO
DATA NASCIMENTO:
02/02/1995
ENDEREÇO RESIDENCIAL:
RUA B
CIDADE/UF - CEP RESIDENCIAL:
RJ/RJ - 20000-000
"""
        regs = parse_texto_tim(texto, "12345678900", "multi.pdf")
        assert len(regs) == 2
        assert regs[0].numero_linha == "5588"
        assert regs[1].numero_linha == "5589"

    def test_registro_quebrado_paginas(self):
        """Simula registro que começa no fim de uma página e continua na outra."""
        texto = """NÚMERO DA LINHA:
5588
TIPO DA LINHA:
PRÉ-PAGO
STATUS ATUAL:
Ativo
DATA CADASTRO:
01/01/2024 00:00:00
NOME:
JOAO
CPF/CNPJ:
123.456.789-00
CONFIDENCIAL
TIM S/A
SEXO:
MASCULINO
DATA NASCIMENTO:
01/01/1990
ENDEREÇO RESIDENCIAL:
RUA A
CIDADE/UF - CEP RESIDENCIAL:
SP/SP - 01000-000
"""
        regs = parse_texto_tim(texto, "12345678900", "quebrado.pdf")
        assert len(regs) == 1
        assert regs[0].sexo == "MASCULINO"
        assert regs[0].endereco_residencial == "RUA A"

    def test_valor_ponto_vazio(self):
        texto = """NÚMERO DA LINHA:
5588
TIPO DA LINHA:
PRÉ-PAGO
STATUS ATUAL:
Ativo
DATA CADASTRO:
01/01/2024 00:00:00
NOME:
JOAO
CPF/CNPJ:
123.456.789-00
SEXO:
MASCULINO
DATA NASCIMENTO:
01/01/1990
TIPO DOCUMENTO:
.
NÚM. DOCUMENTO:
.
NACIONALIDADE:
BRASIL
DATA EMISSÃO:
.
TELEFONE CONTATO:
.
PAÍS EMISSOR:
.
ENDEREÇO RESIDENCIAL:
RUA A
CIDADE/UF - CEP RESIDENCIAL:
SP/SP - 01000-000
"""
        regs = parse_texto_tim(texto, "12345678900", "pontos.pdf")
        assert len(regs) == 1
        assert regs[0].tipo_documento is None
        assert regs[0].num_documento is None
        assert regs[0].nacionalidade == "BRASIL"
        assert regs[0].data_emissao is None
        assert regs[0].telefone_contato is None
        assert regs[0].pais_emissor is None

    def test_bloco_vazio(self):
        texto = """DADOS CADASTRAIS
Não foram encontrados registros referentes à informação solicitada.
DADOS DE HABILITAÇÃO
Não foram encontrados registros referentes à informação solicitada.
"""
        regs = parse_texto_tim(texto, "12345678900", "vazio.pdf")
        assert len(regs) == 0

    def test_registro_incompleto_sem_numero(self):
        texto = """NÚMERO DA LINHA:

TIPO DA LINHA:
PRÉ-PAGO
STATUS ATUAL:
Ativo
DATA CADASTRO:
01/01/2024 00:00:00
NOME:
JOAO
CPF/CNPJ:
123.456.789-00
SEXO:
MASCULINO
DATA NASCIMENTO:
01/01/1990
ENDEREÇO RESIDENCIAL:
RUA A
CIDADE/UF - CEP RESIDENCIAL:
SP/SP - 01000-000
"""
        regs = parse_texto_tim(texto, "12345678900", "incompleto.pdf")
        assert len(regs) == 0

    def test_registro_com_endereco_fatura(self):
        texto = """NÚMERO DA LINHA:
5588
TIPO DA LINHA:
PRÉ-PAGO
STATUS ATUAL:
Ativo
DATA CADASTRO:
01/01/2024 00:00:00
NOME:
JOAO
CPF/CNPJ:
123.456.789-00
SEXO:
MASCULINO
DATA NASCIMENTO:
01/01/1990
ENDEREÇO RESIDENCIAL:
RUA A
CIDADE/UF - CEP RESIDENCIAL:
SP/SP - 01000-000
ENDEREÇO FATURA:
RUA B
CIDADE/UF - CEP FATURA:
RJ/RJ - 20000-000
"""
        regs = parse_texto_tim(texto, "12345678900", "fatura.pdf")
        assert len(regs) == 1
        assert regs[0].endereco_fatura == "RUA B"
        assert regs[0].cidade_uf_cep_fatura == "RJ/RJ - 20000-000"


# ---------------------------------------------------------------------------
# Snapshots
# ---------------------------------------------------------------------------

class TestSnapshots:
    def test_seleciona_mais_recente(self):
        df = pd.DataFrame(
            {
                "numero_linha": ["5588", "5588", "5589"],
                "data_cadastro": [
                    "01/01/2024 10:00:00",
                    "05/01/2024 12:00:00",
                    "02/01/2024 08:00:00",
                ],
                "nome": ["A", "A", "B"],
            }
        )
        df_result = selecionar_snapshots_recentes(df)
        assert len(df_result) == 2
        assert (
            df_result[df_result["numero_linha"] == "5588"]["data_cadastro"].iloc[0]
            == "05/01/2024 12:00:00"
        )

    def test_mesma_data_mantem_um(self):
        df = pd.DataFrame(
            {
                "numero_linha": ["5588", "5588"],
                "data_cadastro": ["01/01/2024 10:00:00", "01/01/2024 10:00:00"],
                "nome": ["A", "A"],
            }
        )
        df_result = selecionar_snapshots_recentes(df)
        assert len(df_result) == 1


# ---------------------------------------------------------------------------
# Deduplicação (via pipeline)
# ---------------------------------------------------------------------------

class TestDeduplicacao:
    def test_remove_duplicados_mesmo_conteudo_arquivo_diferente(self):
        df = pd.DataFrame(
            {
                "cpf_consulta": ["123", "123"],
                "numero_linha": ["5588", "5588"],
                "nome": ["JOAO", "JOAO"],
                "data_cadastro": ["01/01/2024", "01/01/2024"],
                "arquivo_origem": ["a.pdf", "b.pdf"],
            }
        )
        df_dedup = deduplicar_dataframe(df)
        assert len(df_dedup) == 1

    def test_mantem_registros_distintos(self):
        df = pd.DataFrame(
            {
                "cpf_consulta": ["123", "123"],
                "numero_linha": ["5588", "5589"],
                "nome": ["JOAO", "JOAO"],
                "data_cadastro": ["01/01/2024", "01/01/2024"],
                "arquivo_origem": ["a.pdf", "b.pdf"],
            }
        )
        df_dedup = deduplicar_dataframe(df)
        assert len(df_dedup) == 2


# ---------------------------------------------------------------------------
# Excel Estilizado
# ---------------------------------------------------------------------------

class TestEstilizacaoExcel:
    def test_estilo_ativo_verde(self):
        df = pd.DataFrame(
            {
                "status_atual": ["Ativo", "Inativo"],
                "tipo_linha": ["PRÉ-PAGO", "PÓS-PAGO"],
                "nome": ["A", "B"],
            }
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "test.xlsx"
            estilizar_excel(
                df, path, col_situacao="status_atual", col_modalidade="tipo_linha"
            )
            import openpyxl

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            # Header
            assert ws.cell(row=1, column=1).fill.start_color.rgb == "004472C4"
            # Ativo → fill verde
            ativo_fill = ws.cell(row=2, column=1).fill.start_color.rgb
            assert ativo_fill == "00C6EFCE" or ativo_fill == "C6EFCE"

    def test_freeze_panes(self):
        df = pd.DataFrame({"status_atual": ["Ativo"], "nome": ["A"]})
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "test.xlsx"
            estilizar_excel(df, path, col_situacao="status_atual")
            import openpyxl

            wb = openpyxl.load_workbook(path)
            ws = wb.active
            assert ws.freeze_panes == "A2"


# ---------------------------------------------------------------------------
# Geocoding (mockado)
# ---------------------------------------------------------------------------

class TestGeocoding:
    @patch("geocoding.requests.get")
    def test_geocodifica_tim(self, mock_get):
        mock_response = MagicMock()
        mock_response.json.return_value = [{"lat": "-23.5505", "lon": "-46.6333"}]
        mock_response.raise_for_status = MagicMock()
        mock_get.return_value = mock_response

        df = pd.DataFrame(
            {
                "endereco": ["Rua Augusta"],
                "bairro": ["Consolação"],
                "municipio": ["São Paulo"],
                "estado": ["SP"],
                "cep": ["01305-000"],
            }
        )
        with tempfile.TemporaryDirectory() as tmpdir:
            db_path = Path(tmpdir) / "cache.db"
            geocoding.processar_geocoding(df, str(db_path))

        assert df.iloc[0]["latitude"] == -23.5505
        assert df.iloc[0]["longitude"] == -46.6333
        assert "google.com/maps" in df.iloc[0]["google_maps_url"]
        mock_get.assert_called_once()


# ---------------------------------------------------------------------------
# ZIP Security & Processing
# ---------------------------------------------------------------------------

class TestZipSecurity:
    def test_zip_bomb_rejeitado(self):
        """ZIP com ratio de compressão > 100x deve ser rejeitado."""
        z = io.BytesIO()
        with zipfile.ZipFile(z, "w") as zf:
            big = b"0" * (20 * 1024 * 1024)
            zf.writestr("bomb.pdf", big, compress_type=zipfile.ZIP_DEFLATED)
        z.seek(0)

        tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp.write(z.read())
        tmp.close()

        regs = processar_zip(Path(tmp.name))
        assert regs == []

    def test_fake_pdf_rejeitado(self):
        """Arquivo renomeado para .pdf sem magic number %PDF deve ser rejeitado."""
        z = io.BytesIO()
        with zipfile.ZipFile(z, "w") as zf:
            zf.writestr("fake.pdf", b"NOT_A_PDF_AT_ALL")
        z.seek(0)

        tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp.write(z.read())
        tmp.close()

        regs = processar_zip(Path(tmp.name))
        assert regs == []

    def test_path_traversal_sanitizado(self):
        """Nome de arquivo com path traversal deve ser sanitizado nos logs."""
        assert _sanitizar_nome_arquivo("../../../evil.pdf") == "evil.pdf"
        assert _sanitizar_nome_arquivo("pasta/test.pdf") == "test.pdf"
        assert _sanitizar_nome_arquivo("test.pdf") == "test.pdf"

    def test_pdf_muito_grande_rejeitado(self):
        """PDF acima de MAX_PDF_SIZE_BYTES deve ser rejeitado."""
        z = io.BytesIO()
        with zipfile.ZipFile(z, "w") as zf:
            # Cria PDF valido mas grande (usando dummy bytes nao-comprimidos)
            pdf = _criar_pdf_em_memoria("NÚMERO DA LINHA:\n5588\nNOME:\nJOAO")
            # Adiciona padding para ultrapassar limite
            padding = b"\n" * (MAX_PDF_SIZE_BYTES + 1)
            zf.writestr("big.pdf", pdf + padding)
        z.seek(0)

        tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp.write(z.read())
        tmp.close()

        regs = processar_zip(Path(tmp.name))
        assert regs == []

    def test_muitos_pdfs_rejeitado(self):
        """ZIP com mais de MAX_PDFS_PER_ZIP deve ser rejeitado."""
        z = io.BytesIO()
        with zipfile.ZipFile(z, "w") as zf:
            for i in range(MAX_PDFS_PER_ZIP + 1):
                zf.writestr(f"{i}.pdf", b"%PDF1.4 fake")
        z.seek(0)

        tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp.write(z.read())
        tmp.close()

        regs = processar_zip(Path(tmp.name))
        assert regs == []


class TestProcessarZip:
    def test_zip_vazio_sem_pdf(self):
        z = io.BytesIO()
        with zipfile.ZipFile(z, "w") as zf:
            zf.writestr("readme.txt", b"nada")
        z.seek(0)

        tmp = tempfile.NamedTemporaryFile(suffix=".zip", delete=False)
        tmp.write(z.read())
        tmp.close()

        regs = processar_zip(Path(tmp.name))
        assert regs == []

    def test_zip_com_pdf_valido(self):
        """Cria um ZIP com um PDF valido contendo um registro TIM."""
        texto_pdf = """Relatório de Cadastro por (CPF): 12345678901
NÚMERO DA LINHA:
5588
TIPO DA LINHA:
PRÉ-PAGO
STATUS ATUAL:
Ativo
DATA STATUS:
01/01/2024 00:00:00
DATA INÍCIO VÍNCULO:
01/01/2024 00:00:00
DATA CADASTRO:
01/01/2024 00:00:00
DATA FIM VÍNCULO:
.
NOME:
JOAO SILVA
TIPO DO CLIENTE:
CON
CPF/CNPJ:
123.456.789-01
SEXO:
MASCULINO
DATA NASCIMENTO:
01/01/1990
ENDEREÇO RESIDENCIAL:
RUA A
CIDADE/UF - CEP RESIDENCIAL:
SP/SP - 01000-000
"""
        pdf_bytes = _criar_pdf_em_memoria(texto_pdf)
        tmp_zip = _criar_zip_com_pdfs({"pasta/test_12345678901_xxx.pdf": pdf_bytes})

        regs = processar_zip(tmp_zip)
        assert len(regs) == 1
        assert regs[0].numero_linha == "5588"
        assert regs[0].cpf_consulta == "12345678901"
        assert regs[0].nome == "JOAO SILVA"

    def test_cpf_fallback_nome_arquivo(self):
        """Se CPF nao esta no texto do PDF, deve ser extraido do nome do arquivo."""
        texto_pdf = """NÚMERO DA LINHA:
5588
TIPO DA LINHA:
PRÉ-PAGO
STATUS ATUAL:
Ativo
DATA CADASTRO:
01/01/2024 00:00:00
NOME:
JOAO
CPF/CNPJ:
123.456.789-00
SEXO:
MASCULINO
DATA NASCIMENTO:
01/01/1990
ENDEREÇO RESIDENCIAL:
RUA A
CIDADE/UF - CEP RESIDENCIAL:
SP/SP - 01000-000
"""
        pdf_bytes = _criar_pdf_em_memoria(texto_pdf)
        # Nome do arquivo SEM CPF no texto, mas COM CPF no nome
        tmp_zip = _criar_zip_com_pdfs({"rel_98765432100_20240101.pdf": pdf_bytes})

        regs = processar_zip(tmp_zip)
        assert len(regs) == 1
        assert regs[0].cpf_consulta == "98765432100"


# ---------------------------------------------------------------------------
# Path Validation
# ---------------------------------------------------------------------------

class TestPathValidation:
    def test_caminho_absoluto_ok(self):
        _validar_caminho(Path("C:\\teste"), "input")
        # No Windows, /tmp/teste nao e absoluto; usar caminho UNC ou C:
        import sys
        if sys.platform == "win32":
            _validar_caminho(Path("D:\\teste"), "output")
        else:
            _validar_caminho(Path("/tmp/teste"), "output")

    def test_path_traversal_rejeitado(self):
        with pytest.raises(ValueError):
            _validar_caminho(Path("C:\\teste\\..\\etc"), "input")

    def test_caminho_relativo_rejeitado(self):
        with pytest.raises(ValueError):
            _validar_caminho(Path("teste"), "input")


# ---------------------------------------------------------------------------
# PDF Extraction
# ---------------------------------------------------------------------------

class TestPdfExtraction:
    def test_extrair_texto_pdf_valido(self):
        texto = "NÚMERO DA LINHA:\n5588\nNOME:\nJOAO"
        pdf_bytes = _criar_pdf_em_memoria(texto)
        extraido = extrair_texto_pdf_de_bytes(pdf_bytes)
        assert "NÚMERO DA LINHA:" in extraido
        assert "5588" in extraido

    def test_magic_number_invalido_rejeita(self):
        extraido = extrair_texto_pdf_de_bytes(b"NOT_A_PDF")
        assert extraido == ""

    def test_pdf_bytes_vazio(self):
        extraido = extrair_texto_pdf_de_bytes(b"")
        assert extraido == ""


# ---------------------------------------------------------------------------
# Main / CLI Integration
# ---------------------------------------------------------------------------

class TestMain:
    @patch("parse_tim.processar_zip")
    @patch("parse_tim.geocoding.processar_geocoding")
    def test_main_completo_sem_geocode(self, mock_geo, mock_processar_zip):
        from parse_tim import main

        reg = RegistroTim(
            cpf_consulta="123",
            numero_linha="5588",
            tipo_linha="PRÉ-PAGO",
            status_atual="Ativo",
            data_status="01/01/2024",
            data_inicio_vinculo="01/01/2024",
            data_cadastro="01/01/2024 10:00:00",
            data_fim_vinculo=None,
            nome="JOAO",
            tipo_cliente="CON",
            cpf_cnpj="123.456.789-00",
            sexo="MASCULINO",
            tipo_documento=None,
            data_nascimento="01/01/1990",
            num_documento=None,
            nacionalidade="BRASIL",
            data_emissao=None,
            telefone_contato=None,
            pais_emissor=None,
            endereco_residencial="RUA A",
            cidade_uf_cep_residencial="SP/SP - 01000-000",
            endereco_fatura=None,
            cidade_uf_cep_fatura=None,
            arquivo_origem="teste.pdf",
        )
        mock_processar_zip.return_value = [reg]

        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "input"
            output_dir = Path(tmpdir) / "output"
            input_dir.mkdir()
            # Cria um ZIP dummy para ser encontrado pelo glob
            (input_dir / "dummy.zip").write_bytes(b"PK")

            with patch("sys.argv", [
                "parse_tim.py",
                "--input", str(input_dir),
                "--output", str(output_dir),
                "--no-geocode",
            ]):
                main()

            assert (output_dir / "dados_cadastrais_tim.xlsx").exists()
            assert (output_dir / "dados_cadastrais.db").exists()
            mock_geo.assert_not_called()

    @patch("parse_tim.processar_zip")
    def test_main_historico(self, mock_processar_zip):
        from parse_tim import main

        regs = [
            RegistroTim(
                cpf_consulta="123", numero_linha="5588", tipo_linha="PRÉ-PAGO",
                status_atual="Ativo", data_status="01/01/2024", data_inicio_vinculo="01/01/2024",
                data_cadastro="01/01/2024 10:00:00", data_fim_vinculo=None, nome="JOAO",
                tipo_cliente="CON", cpf_cnpj="123.456.789-00", sexo="MASCULINO",
                tipo_documento=None, data_nascimento="01/01/1990", num_documento=None,
                nacionalidade="BRASIL", data_emissao=None, telefone_contato=None,
                pais_emissor=None, endereco_residencial="RUA A",
                cidade_uf_cep_residencial="SP/SP - 01000-000", endereco_fatura=None,
                cidade_uf_cep_fatura=None, arquivo_origem="a.pdf",
            ),
            RegistroTim(
                cpf_consulta="123", numero_linha="5588", tipo_linha="PRÉ-PAGO",
                status_atual="Ativo", data_status="01/01/2024", data_inicio_vinculo="01/01/2024",
                data_cadastro="02/01/2024 10:00:00", data_fim_vinculo=None, nome="JOAO",
                tipo_cliente="CON", cpf_cnpj="123.456.789-00", sexo="MASCULINO",
                tipo_documento=None, data_nascimento="01/01/1990", num_documento=None,
                nacionalidade="BRASIL", data_emissao=None, telefone_contato=None,
                pais_emissor=None, endereco_residencial="RUA A",
                cidade_uf_cep_residencial="SP/SP - 01000-000", endereco_fatura=None,
                cidade_uf_cep_fatura=None, arquivo_origem="b.pdf",
            ),
        ]
        mock_processar_zip.return_value = regs

        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "input"
            output_dir = Path(tmpdir) / "output"
            input_dir.mkdir()
            (input_dir / "dummy.zip").write_bytes(b"PK")

            with patch("sys.argv", [
                "parse_tim.py",
                "--input", str(input_dir),
                "--output", str(output_dir),
                "--no-geocode",
                "--historico",
            ]):
                main()

            import sqlite3
            conn = sqlite3.connect(output_dir / "dados_cadastrais.db")
            cursor = conn.execute("SELECT count(*) FROM linhas_tim")
            count = cursor.fetchone()[0]
            conn.close()
            assert count == 2  # modo historico mantem os 2 snapshots

    def test_main_nenhum_zip(self):
        from parse_tim import main

        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "input"
            output_dir = Path(tmpdir) / "output"
            input_dir.mkdir()

            with patch("sys.argv", [
                "parse_tim.py",
                "--input", str(input_dir),
                "--output", str(output_dir),
                "--no-geocode",
            ]):
                main()  # deve retornar sem erro após log

    @patch("parse_tim.processar_zip")
    def test_main_nenhum_registro(self, mock_processar_zip):
        from parse_tim import main

        mock_processar_zip.return_value = []

        with tempfile.TemporaryDirectory() as tmpdir:
            input_dir = Path(tmpdir) / "input"
            output_dir = Path(tmpdir) / "output"
            input_dir.mkdir()
            (input_dir / "dummy.zip").write_bytes(b"PK")

            with patch("sys.argv", [
                "parse_tim.py",
                "--input", str(input_dir),
                "--output", str(output_dir),
                "--no-geocode",
            ]):
                main()  # deve retornar sem erro após log


# ---------------------------------------------------------------------------
# Dataclass RegistroTim
# ---------------------------------------------------------------------------

class TestRegistroTim:
    def test_criacao_completa(self):
        r = RegistroTim(
            cpf_consulta="123",
            numero_linha="5588",
            tipo_linha="PRÉ-PAGO",
            status_atual="Ativo",
            data_status="01/01/2024",
            data_inicio_vinculo="01/01/2024",
            data_cadastro="01/01/2024",
            data_fim_vinculo=None,
            nome="JOAO",
            tipo_cliente="CON",
            cpf_cnpj="123.456.789-00",
            sexo="MASCULINO",
            tipo_documento=None,
            data_nascimento="01/01/1990",
            num_documento=None,
            nacionalidade="BRASIL",
            data_emissao=None,
            telefone_contato=None,
            pais_emissor=None,
            endereco_residencial="RUA A",
            cidade_uf_cep_residencial="SP/SP - 01000-000",
            endereco_fatura=None,
            cidade_uf_cep_fatura=None,
            arquivo_origem="teste.pdf",
        )
        assert r.numero_linha == "5588"
        assert r.data_fim_vinculo is None

    def test_asdict(self):
        r = RegistroTim(
            cpf_consulta="123",
            numero_linha="5588",
            tipo_linha="PRÉ-PAGO",
            status_atual="Ativo",
            data_status="01/01/2024",
            data_inicio_vinculo="01/01/2024",
            data_cadastro="01/01/2024",
            data_fim_vinculo=None,
            nome="JOAO",
            tipo_cliente="CON",
            cpf_cnpj="123.456.789-00",
            sexo="MASCULINO",
            tipo_documento=None,
            data_nascimento="01/01/1990",
            num_documento=None,
            nacionalidade="BRASIL",
            data_emissao=None,
            telefone_contato=None,
            pais_emissor=None,
            endereco_residencial="RUA A",
            cidade_uf_cep_residencial="SP/SP - 01000-000",
            endereco_fatura=None,
            cidade_uf_cep_fatura=None,
            arquivo_origem="teste.pdf",
        )
        d = r.__dict__
        assert d["numero_linha"] == "5588"
