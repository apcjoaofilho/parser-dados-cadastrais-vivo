import sys
from pathlib import Path

_parser_dir = Path(__file__).resolve().parent.parent.parent / 'vivo-parser'
sys.path.insert(0, str(_parser_dir))
_shared_dir = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_shared_dir))

import sqlite3
import tempfile
from pathlib import Path
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

from parse_vivo import (
    deduplicar_dataframe, estilizar_excel, ordenar_dataframe,
    parse_arquivo, parse_registro, salvar_sqlite, _validar_path
)
import geocoding


# ============================================================================
# Fixtures
# ============================================================================

@pytest.fixture
def sample_txt_content():
    """Conteúdo de um arquivo .txt de exemplo com 2 registros."""
    return """* ---------------------------------------------------------------------------- *
*                           PARÂMETRO(S) DE CONSULTA                           *
*                                                                              *
* CPF: 123.456.789-00                                                          *
* ---------------------------------------------------------------------------- *
* NÚMERO DA LINHA:.............................................(11) 98765-4321 *
* CLIENTE:.....................................................JOAO DA SILVA *
* CPF:..........................................................123.456.789-00 *
* ENDEREÇO:........................................Rua A, 123, Centro        *
* BAIRRO:..............................................................CENTRO *
* CEP:..............................................................01.000-000 *
* MUNICÍPIO:...........................................................SAO PAULO *
* ESTADO:..................................................................SP *
* MODALIDADE:..............................................................PRE *
* SITUAÇÃO:..............................................................ATIVO *
* DATA HABILITAÇÃO:.................................................01/01/2024 *
* ............................................................................ *
* NÚMERO DA LINHA:.............................................(11) 91234-5678 *
* CLIENTE:.....................................................MARIA OLIVEIRA *
* CPF:..........................................................123.456.789-00 *
* ENDEREÇO:........................................Rua B, 456, Jardins      *
* BAIRRO:.............................................................JARDINS *
* CEP:..............................................................02.000-000 *
* MUNICÍPIO:...........................................................SAO PAULO *
* ESTADO:..................................................................SP *
* MODALIDADE:..............................................................POS *
* SITUAÇÃO:............................................................INATIVO *
* DATA HABILITAÇÃO:.................................................15/06/2023 *
* DATA RESCISÃO:....................................................20/12/2023 *
* ............................................................................ *
* ---------------------------------------------------------------------------- *
"""


@pytest.fixture
def sample_registro_df():
    """DataFrame com registros de exemplo para testes."""
    return pd.DataFrame({
        'cpf_consulta': ['123', '123', '456', '456'],
        'numero_linha': ['(11) 1111-1111', '(11) 2222-2222', '(11) 3333-3333', '(11) 4444-4444'],
        'cliente': ['JOAO', 'JOAO', 'MARIA', 'MARIA'],
        'cpf': ['123', '123', '456', '456'],
        'endereco': ['Rua A', 'Rua A', 'Rua B', 'Rua C'],
        'bairro': ['Centro', 'Centro', 'Jardins', 'Jardins'],
        'cep': ['01000-000', '01000-000', '02000-000', '02000-000'],
        'municipio': ['SP', 'SP', 'SP', 'SP'],
        'estado': ['SP', 'SP', 'SP', 'SP'],
        'modalidade': ['PRE', 'PRE', 'POS', 'POS'],
        'situacao': ['ATIVO', 'ATIVO', 'INATIVO', 'ATIVO'],
        'data_habilitacao': ['01/01/2024', '01/01/2024', '15/03/2023', '20/06/2024'],
        'data_rescisao': [None, None, '20/12/2023', None],
        'arquivo_origem': ['a.txt', 'b.txt', 'c.txt', 'd.txt'],
    })


# ============================================================================
# Testes de Integração (End-to-End)
# ============================================================================

class TestIntegracao:
    def test_pipeline_completo_arquivo_txt(self, sample_txt_content, tmp_path):
        """Testa o pipeline completo: .txt → DataFrame → Excel → SQLite."""
        # Cria arquivo de entrada
        input_dir = tmp_path / "input"
        input_dir.mkdir()
        txt_file = input_dir / "teste.txt"
        txt_file.write_text(sample_txt_content, encoding='utf-8')
        
        # Parse
        registros = parse_arquivo(txt_file)
        assert len(registros) == 2
        
        # Cria DataFrame
        from dataclasses import asdict
        df_raw = pd.DataFrame([asdict(r) for r in registros])
        assert len(df_raw) == 2
        
        # Deduplica
        df_dedup = deduplicar_dataframe(df_raw)
        assert len(df_dedup) == 2  # registros são distintos
        
        # Ordena
        df_dedup = ordenar_dataframe(df_dedup)
        assert df_dedup.iloc[0]['situacao'] == 'ATIVO'
        
        # Gera Excel
        output_dir = tmp_path / "output"
        output_dir.mkdir()
        excel_path = output_dir / "teste.xlsx"
        estilizar_excel(df_dedup, excel_path)
        assert excel_path.exists()
        
        # Verifica Excel
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.max_row == 3  # header + 2 registros
        assert ws.max_column == 14
        assert ws.freeze_panes == 'A2'
        
        # Gera SQLite
        db_path = output_dir / "teste.db"
        salvar_sqlite(df_raw, df_dedup, db_path, tabela_raw='linhas_raw', tabela_dedup='linhas')
        assert db_path.exists()
        
        # Verifica SQLite
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM linhas_raw')
        assert cursor.fetchone()[0] == 2
        cursor.execute('SELECT COUNT(*) FROM linhas')
        assert cursor.fetchone()[0] == 2
        cursor.execute('SELECT * FROM processamento_log')
        log = cursor.fetchone()
        assert log[2] == 2  # total_registros
        assert log[3] == 2  # total_registros_dedup
        conn.close()
    
    def test_arquivo_vazio_retorna_lista_vazia(self, tmp_path):
        """Arquivo sem registros deve retornar lista vazia."""
        txt_file = tmp_path / "vazio.txt"
        txt_file.write_text("* Nenhum Dado Foi Encontrado para Essa Pesquisa *", encoding='utf-8')
        registros = parse_arquivo(txt_file)
        assert registros == []
    
    def test_multiplos_arquivos_com_duplicatas(self, tmp_path):
        """Testa deduplicação quando o mesmo registro aparece em arquivos diferentes."""
        # Cria DataFrame com registros duplicados (mesmo conteúdo, arquivo diferente)
        df = pd.DataFrame({
            'cpf_consulta': ['123', '123'],
            'numero_linha': ['(11) 1111-1111', '(11) 1111-1111'],
            'cliente': ['JOAO', 'JOAO'],
            'cpf': ['123', '123'],
            'endereco': ['Rua A', 'Rua A'],
            'bairro': ['Centro', 'Centro'],
            'cep': ['01000-000', '01000-000'],
            'municipio': ['SP', 'SP'],
            'estado': ['SP', 'SP'],
            'modalidade': ['PRE', 'PRE'],
            'situacao': ['ATIVO', 'ATIVO'],
            'data_habilitacao': ['01/01/2024', '01/01/2024'],
            'data_rescisao': [None, None],
            'arquivo_origem': ['a.txt', 'b.txt'],
        })
        df_dedup = deduplicar_dataframe(df)
        
        # Deve remover 1 duplicata
        assert len(df_dedup) == 1
        assert 'a.txt' in df_dedup['arquivo_origem'].values
        assert 'b.txt' not in df_dedup['arquivo_origem'].values  # keep='first'


# ============================================================================
# Testes de Pipeline Detalhados
# ============================================================================

class TestPipelineDetalhado:
    def test_ordenacao_por_data_descendente(self):
        """ATIVO mais recente deve vir primeiro, depois ATIVO mais antigo, depois INATIVO."""
        df = pd.DataFrame({
            'cpf_consulta': ['1', '2', '3', '4'],
            'numero_linha': ['A', 'B', 'C', 'D'],
            'cliente': ['X', 'Y', 'Z', 'W'],
            'cpf': ['1', '2', '3', '4'],
            'endereco': ['Rua', 'Rua', 'Rua', 'Rua'],
            'bairro': ['Centro', 'Centro', 'Centro', 'Centro'],
            'cep': ['00000-000', '00000-000', '00000-000', '00000-000'],
            'municipio': ['SP', 'SP', 'SP', 'SP'],
            'estado': ['SP', 'SP', 'SP', 'SP'],
            'modalidade': ['PRE', 'PRE', 'PRE', 'PRE'],
            'situacao': ['INATIVO', 'ATIVO', 'ATIVO', 'INATIVO'],
            'data_habilitacao': ['01/01/2022', '15/03/2025', '10/01/2024', '20/06/2023'],
            'data_rescisao': [None, None, None, None],
            'arquivo_origem': ['a.txt', 'b.txt', 'c.txt', 'd.txt'],
        })
        df_ord = ordenar_dataframe(df)
        
        # ATIVO primeiro, mais recente primeiro
        assert df_ord.iloc[0]['situacao'] == 'ATIVO'
        assert df_ord.iloc[0]['data_habilitacao'] == '15/03/2025'
        assert df_ord.iloc[1]['situacao'] == 'ATIVO'
        assert df_ord.iloc[1]['data_habilitacao'] == '10/01/2024'
        # INATIVO por último
        assert df_ord.iloc[2]['situacao'] == 'INATIVO'
        assert df_ord.iloc[3]['situacao'] == 'INATIVO'
    
    def test_excel_com_colunas_geocoding(self, tmp_path):
        """Excel deve incluir colunas latitude, longitude, google_maps_url quando presentes."""
        df = pd.DataFrame({
            'situacao': ['ATIVO'],
            'modalidade': ['PRE'],
            'cliente': ['A'],
            'latitude': [-23.5],
            'longitude': [-46.6],
            'google_maps_url': ['https://maps.google.com/?q=-23.5,-46.6'],
        })
        excel_path = tmp_path / 'teste.xlsx'
        estilizar_excel(df, excel_path)
        
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        assert ws.max_column == 6
        # Verifica que o header está correto
        assert ws.cell(row=1, column=4).value == 'latitude'
        assert ws.cell(row=1, column=5).value == 'longitude'
        assert ws.cell(row=1, column=6).value == 'google_maps_url'


# ============================================================================
# Testes de Geocodificação com Mock
# ============================================================================

class TestGeocodingMock:
    @patch('geocoding.requests.get')
    def test_cache_hit_nao_faz_requisicao(self, mock_get, tmp_path):
        """Se o endereço já está no cache, não deve fazer requisição HTTP."""
        mock_response = MagicMock()
        mock_response.json.return_value = [{'lat': '-23.5505', 'lon': '-46.6333'}]
        mock_response.raise_for_status = MagicMock()
        mock_get.return_value = mock_response
        
        df = pd.DataFrame({
            'endereco': ['Rua Augusta'],
            'bairro': ['Consolacao'],
            'municipio': ['Sao Paulo'],
            'estado': ['SP'],
            'cep': ['01305-000'],
        })
        db_path = tmp_path / 'cache.db'
        
        # Primeira chamada - deve fazer requisição
        geocoding.processar_geocoding(df, str(db_path))
        assert mock_get.call_count == 1
        assert df.iloc[0]['latitude'] == -23.5505
        
        # Segunda chamada com mesmo endereço - deve usar cache
        df2 = pd.DataFrame({
            'endereco': ['Rua Augusta'],
            'bairro': ['Consolacao'],
            'municipio': ['Sao Paulo'],
            'estado': ['SP'],
            'cep': ['01305-000'],
        })
        geocoding.processar_geocoding(df2, str(db_path))
        # Não deve ter feito mais requisições
        assert mock_get.call_count == 1
        assert df2.iloc[0]['latitude'] == -23.5505
    
    @patch('geocoding.requests.get')
    def test_falha_geocodificacao_registra_none(self, mock_get, tmp_path):
        """Se a geocodificação falhar, deve registrar None sem quebrar."""
        mock_get.side_effect = Exception('Timeout')
        
        df = pd.DataFrame({
            'endereco': ['Rua Inexistente'],
            'bairro': ['Bairro X'],
            'municipio': ['Cidade Y'],
            'estado': ['ZZ'],
            'cep': ['00000-000'],
        })
        db_path = tmp_path / 'cache.db'
        geocoding.processar_geocoding(df, str(db_path))
        
        assert pd.isna(df.iloc[0]['latitude'])
        assert pd.isna(df.iloc[0]['longitude'])
        assert pd.isna(df.iloc[0]['google_maps_url'])
    
    @patch('geocoding.requests.get')
    def test_variacoes_endereco_cascade(self, mock_get, tmp_path):
        """Deve tentar variações do endereço quando a primeira falha."""
        # Primeira chamada retorna vazio, segunda retorna coordenadas
        mock_get.side_effect = [
            MagicMock(json=lambda: [], raise_for_status=lambda: None),  # vazio
            MagicMock(json=lambda: [{'lat': '-23.5', 'lon': '-46.6'}], raise_for_status=lambda: None),  # sucesso
        ]
        
        df = pd.DataFrame({
            'endereco': ['Rua Augusta 123 Apto 45'],
            'bairro': ['Consolacao'],
            'municipio': ['Sao Paulo'],
            'estado': ['SP'],
            'cep': ['01305-000'],
        })
        db_path = tmp_path / 'cache.db'
        geocoding.processar_geocoding(df, str(db_path))
        
        # Deve ter tentado pelo menos 2 variações
        assert mock_get.call_count >= 2
        assert df.iloc[0]['latitude'] == -23.5


# ============================================================================
# Testes de Segurança
# ============================================================================

class TestSeguranca:
    def test_path_traversal_rejeitado(self):
        """Paths com .. devem ser rejeitados."""
        with pytest.raises(ValueError, match='inseguro'):
            _validar_path('../../etc/passwd', deve_existir=False)
    
    def test_path_relativo_rejeitado(self):
        """Paths relativos devem ser rejeitados."""
        with pytest.raises(ValueError, match='inseguro'):
            _validar_path('relative/path', deve_existir=False)
    
    def test_path_absoluto_seguro(self, tmp_path):
        """Paths absolutos válidos devem ser aceitos."""
        caminho = _validar_path(str(tmp_path), deve_existir=True)
        assert caminho.is_absolute()
    
    def test_path_nao_existente_rejeitado_quando_deve_existir(self, tmp_path):
        """Paths que não existem devem ser rejeitados quando deve_existir=True."""
        with pytest.raises(FileNotFoundError):
            _validar_path(str(tmp_path / 'nao_existe'), deve_existir=True)
    
    def test_path_nao_existente_aceito_quando_nao_precisa_existir(self, tmp_path):
        """Paths que não existem devem ser aceitos quando deve_existir=False."""
        caminho = _validar_path(str(tmp_path / 'nao_existe'), deve_existir=False)
        assert caminho.is_absolute()


# ============================================================================
# Testes de Registro Completo
# ============================================================================

class TestRegistroCompleto:
    def test_registro_com_endereco_multilinha(self):
        """Endereço que ocupa múltiplas linhas deve ser concatenado."""
        linhas = [
            '* NÚMERO DA LINHA:.............................................(21) 99826-1047 *',
            '* CLIENTE:.......................................ANTONIA BRUNA DO AMARAL SOUSA *',
            '* CPF:..........................................................067.506.713-81 *',
            '* ENDEREÇO:Rua R DR ARY CASTELO BRANCO UCHOA 209 AP 5 CON NAIR 209 209 AP 5 CON ',
            '* NAIR 209                                                                     *',
            '* BAIRRO:..........................................................REIS VELOSO *',
            '* CEP:..............................................................64.204-010 *',
            '* MUNICÍPIO:..........................................................PARNAIBA *',
            '* ESTADO:...................................................................PI *',
            '* MODALIDADE:..............................................................PRE *',
            '* SITUAÇÃO:............................................................INATIVO *',
            '* DATA HABILITAÇÃO:.................................................16/01/2021 *',
            '* DATA RESCISÃO:....................................................23/11/2021 *',
            '* ............................................................................ *',
        ]
        reg, nxt = parse_registro(linhas, 0, '067.506.713-81', 'teste.txt')
        assert reg is not None
        assert 'ARY CASTELO BRANCO UCHOA 209 AP 5 CON NAIR 209 209 AP 5 CON NAIR 209' in reg.endereco
    
    def test_bloco_sem_registros(self):
        """Arquivo com 'Nenhum Dado Foi Encontrado' deve retornar lista vazia."""
        linhas = [
            '* ---------------------------------------------------------------------------- *',
            '*                           PARÂMETRO(S) DE CONSULTA                           *',
            '*                                                                              *',
            '* CPF: 069.552.203-57                                                          *',
            '* ---------------------------------------------------------------------------- *',
            '* Nenhum Dado Foi Encontrado para Essa Pesquisa............................... *',
            '* ............................................................................ *',
        ]
        from parse_vivo import parse_arquivo_from_lines
        regs = parse_arquivo_from_lines(linhas, 'vazio.txt')
        assert len(regs) == 0
