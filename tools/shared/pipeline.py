#!/usr/bin/env python3
"""
Pipeline genérico para parsers de dados cadastrais.
Funções compartilhadas: deduplicação, ordenação, estilização Excel e SQLite.
"""

import logging
import sqlite3
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

import geocoding

logger = logging.getLogger(__name__)


def _parse_data(valor: str, fmt: str = "%d/%m/%Y") -> pd.Timestamp | None:
    """Converte string para Timestamp com formato customizável."""
    try:
        return pd.to_datetime(valor, format=fmt, errors="raise")
    except Exception:
        return None


def deduplicar_dataframe(df: pd.DataFrame, ignore_cols: list[str] | None = None) -> pd.DataFrame:
    """Remove duplicatas ignorando colunas especificadas (padrão: arquivo_origem)."""
    if ignore_cols is None:
        ignore_cols = ["arquivo_origem"]
    cols = [c for c in df.columns if c not in ignore_cols]
    df_dedup = df.drop_duplicates(subset=cols, keep="first").reset_index(drop=True)
    removidos = len(df) - len(df_dedup)
    if removidos:
        logger.info("Deduplicação: %d registros removidos (%d → %d)", removidos, len(df), len(df_dedup))
    return df_dedup


def ordenar_dataframe(
    df: pd.DataFrame,
    col_situacao: str = "situacao",
    col_data: str = "data_habilitacao",
    data_format: str = "%d/%m/%Y",
) -> pd.DataFrame:
    """Ordena por situação (ATIVO/Ativo primeiro) e data descendente."""
    df = df.copy()
    df["_data_dt"] = df[col_data].apply(lambda x: _parse_data(str(x), data_format))
    df["_situacao_ordem"] = df[col_situacao].apply(
        lambda x: 0 if str(x).strip().upper() == "ATIVO" else 1
    )
    df = df.sort_values(by=["_situacao_ordem", "_data_dt"], ascending=[True, False]).reset_index(drop=True)
    df = df.drop(columns=["_situacao_ordem", "_data_dt"])
    return df


def estilizar_excel(
    df: pd.DataFrame,
    excel_path: Path,
    col_situacao: str = "situacao",
    col_modalidade: str = "modalidade",
    larguras: dict[str, int] | None = None,
) -> None:
    """Gera arquivo Excel formatado com estilos visuais (zebra, destaque ATIVO, negrito POS, mobile-friendly)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Dados Cadastrais"

    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    fill_white = PatternFill(fill_type=None)
    fill_gray = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    fill_green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    font_green = Font(color="006100")
    font_bold = Font(bold=True)
    alignment_wrap = Alignment(wrap_text=True, vertical="top")

    # Determinar índice das colunas relevantes
    col_situacao_idx = df.columns.get_loc(col_situacao) if col_situacao in df.columns else None
    col_modalidade_idx = df.columns.get_loc(col_modalidade) if col_modalidade in df.columns else None

    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                situacao_val = (
                    str(df.iloc[r_idx - 2][col_situacao]).strip().upper()
                    if col_situacao_idx is not None
                    else ""
                )
                modalidade_val = (
                    str(df.iloc[r_idx - 2][col_modalidade]).strip().upper()
                    if col_modalidade_idx is not None
                    else ""
                )

                if situacao_val == "ATIVO":
                    cell.fill = fill_green
                    cell.font = font_green
                else:
                    cell.fill = fill_gray if (r_idx % 2 == 0) else fill_white

                if modalidade_val == "POS":
                    # Preservar cor da fonte se já for verde (ATIVO)
                    cor_atual = cell.font.color
                    cell.font = Font(bold=True, color=cor_atual)

                cell.alignment = alignment_wrap

    # Larguras fixas mobile-friendly
    larguras_padrao = {
        "cpf_consulta": 16,
        "numero_linha": 18,
        "cliente": 30,
        "cpf": 16,
        "endereco": 40,
        "bairro": 18,
        "cep": 12,
        "municipio": 18,
        "estado": 8,
        "modalidade": 12,
        "situacao": 12,
        "data_habilitacao": 16,
        "data_rescisao": 16,
        "arquivo_origem": 22,
        "latitude": 14,
        "longitude": 14,
        "google_maps_url": 45,
        "nome": 30,
        "tipo_linha": 14,
        "status_atual": 14,
        "data_status": 18,
        "data_inicio_vinculo": 18,
        "data_cadastro": 18,
        "data_fim_vinculo": 18,
        "tipo_cliente": 14,
        "cpf_cnpj": 18,
        "sexo": 10,
        "tipo_documento": 16,
        "data_nascimento": 16,
        "num_documento": 16,
        "nacionalidade": 14,
        "data_emissao": 16,
        "telefone_contato": 18,
        "pais_emissor": 14,
        "endereco_residencial": 40,
        "cidade_uf_cep_residencial": 28,
        "endereco_fatura": 40,
        "cidade_uf_cep_fatura": 28,
    }
    if larguras:
        larguras_padrao.update(larguras)

    for col_name in df.columns:
        col_letter = ws.cell(row=1, column=df.columns.get_loc(col_name) + 1).column_letter
        ws.column_dimensions[col_letter].width = larguras_padrao.get(col_name, 15)

    # Congelar primeira linha
    ws.freeze_panes = "A2"

    wb.save(excel_path)
    logger.info("Excel estilizado salvo em %s", excel_path)


def salvar_sqlite(
    df_raw: pd.DataFrame,
    df_dedup: pd.DataFrame,
    db_path: Path,
    tabela_raw: str = "linhas_raw",
    tabela_dedup: str = "linhas",
) -> None:
    """Salva DataFrames raw e deduplicado em SQLite, além de log de processamento."""
    conn = sqlite3.connect(db_path)
    df_raw.to_sql(tabela_raw, conn, if_exists="replace", index=False)
    df_dedup.to_sql(tabela_dedup, conn, if_exists="replace", index=False)

    # Garantir tabela de cache exista (mesmo que geocoding não rode)
    geocoding.criar_tabela_cache(conn)

    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS processamento_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data_processamento TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            total_registros INTEGER,
            total_registros_dedup INTEGER,
            arquivo_db TEXT
        )
        """
    )
    conn.execute(
        "INSERT INTO processamento_log (total_registros, total_registros_dedup, arquivo_db) VALUES (?, ?, ?)",
        (len(df_raw), len(df_dedup), str(db_path)),
    )
    conn.commit()
    conn.close()
    logger.info("SQLite salvo em %s (raw=%d, dedup=%d)", db_path, len(df_raw), len(df_dedup))
